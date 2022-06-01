import os
import sys
import json
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
try:
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter
from convertJsonToExcelInterface import ConvertJsonToExcelInterface
import color_table
reload(sys)
sys.setdefaultencoding('utf8')

class GenerateToExcel(ConvertJsonToExcelInterface):
    """docstring for GenerateToExcel"""
    def __init__(self,p_release_name):
        super(GenerateToExcel, self).__init__(p_release_name)
    def WriteColumnHeader(self):
        list_column = self.GetColumnHeader()
        # print list_column
        list_init_col = self.init_header_column.keys()
        list_addition_col = self.addition_header_column.keys()

        column_temp = 0
        for index in range(len(list_column)):
            column_temp += 1
            # self.column_number.append(column_temp)
            self.string_to_cell(list_column[index], column_temp, 1, self.excel_header_row_color)

            if list_column[index] in list_init_col:
                len_sub_col = len(self.init_header_column[list_column[index]]["sub_column"])
                if not len_sub_col:
                    self.list_merge_cell.append(self.RowColToCoordinate(column_temp, 1) + ':' + self.RowColToCoordinate(column_temp, 2))

                if "LIST_DERIVATIVE" in self.init_header_column[list_column[index]]["sub_column"]:
                    self.list_merge_cell.append(self.RowColToCoordinate(column_temp, 1) + ':' + self.RowColToCoordinate(column_temp + len(self.list_derivative) - 1, 1))
                    for count in range(len(self.list_derivative)):
                        self.string_to_cell(self.list_derivative[count], column_temp + count, 2, self.excel_header_row_color)
                    column_temp += len(self.list_derivative) - 1

            if list_column[index] in list_addition_col:
                len_sub_col = len(self.addition_header_column[list_column[index]]["sub_column"])
                if not len_sub_col:
                    self.list_merge_cell.append(self.RowColToCoordinate(column_temp, 1) + ':' + self.RowColToCoordinate(column_temp, 2))
                else:
                    self.list_merge_cell.append(self.RowColToCoordinate(column_temp, 1) + ':' + self.RowColToCoordinate(column_temp + len_sub_col - 1, 1))
                    for count in range(len_sub_col):
                        self.string_to_cell(self.addition_header_column[list_column[index]]["sub_column"][count], column_temp + count, 2, self.excel_header_row_color)
                    column_temp += len_sub_col - 1


    def WriteDiffData(self, f_prev_data, f_cur_data, f_col_header, f_col_index, f_row, f_group_color):
        cell_val = f_cur_data
        is_increase = False
        is_change = False

        if isinstance(cell_val, int) or isinstance(cell_val, float):
            cell_val = int(round(cell_val))
            if isinstance(f_prev_data, str) or isinstance(f_prev_data, unicode):
                cell_val = str(cell_val) + "(+" + str(cell_val) + ")"
                is_increase = True
                is_change = True
            else:
                if cell_val > f_prev_data:
                    is_increase = True
                    is_change = True
                    cell_val = str(cell_val) + "(+" + str(cell_val - f_prev_data) + ")"
                if cell_val < f_prev_data:
                    is_increase = False
                    is_change = True
                    cell_val = str(cell_val) + "(" + str(cell_val - f_prev_data) + ")"

        self.string_to_cell(cell_val, self.col_info[f_col_header] + f_col_index, f_row, f_group_color)

        if "Failed tests number on ATE" not in f_col_header:
            if is_change is True:
                if is_increase is True:
                    color_cell = color_table.color_list[str(''.join(self.init_header_column[f_col_header]["increase_color"]))]
                else:
                    color_cell = color_table.color_list[str(''.join(self.init_header_column[f_col_header]["decrease_color"]))]

                self.work_sheet.cell(row=f_row, column=self.col_info[f_col_header] + f_col_index).font = Font(bold=True, color=color_cell)

    def ParseJsonToExcel(self):
        # print self.column_number
        list_module_has_no_json = []

        if not self.report_info['current_report_json']:
            self.current_report_json += self.GetLatestReportFolder(self.current_report_json)

        if not self.report_info['previos_report_json']:
            self.previos_report_json += self.GetLatestReportFolder(self.previos_report_json)

        print "Current report folder: ", self.current_report_json
        print "Previous report folder: ", self.previos_report_json
        json_path_template = os.path.join(self.current_report_json, "MODULE.json")
        json_path_previous_template = os.path.join(self.previos_report_json, "MODULE.json")
        row_count = self.number_row_header + 1

        self.col_info = dict([(key, {}) for key in self.init_header_column.keys()])

        for key in self.init_header_column.keys():
            self.col_info[key] = self.GetColumnNumber(key)

        for group in self.list_group:
            start_group_row = row_count
            group_color = color_table.color_list[str(self.group_color[self.list_group.index(group)])]
            self.string_to_cell(group, self.col_info["Group"], row_count, group_color)
            number_of_module = 0
            list_member = self.module_allocate[group].keys()
            # list_member.sort()
            temp_count = 0
            for mem_count in range(len(list_member)):
                self.string_to_cell(list_member[mem_count], self.col_info["Owner"], row_count, group_color, 'left')
                list_module = self.module_allocate[group][list_member[mem_count]]
                number_of_module += len(list_module)
                if len(list_module) > 1:
                    self.list_merge_cell.append(self.RowColToCoordinate(self.col_info["Owner"], row_count) + ':' + self.RowColToCoordinate(self.col_info["Owner"], row_count + len(list_module) - 1))
                for module_count in range(len(list_module)):
                    self.string_to_cell(list_module[module_count], self.col_info["Module"], row_count, group_color, 'left')

                    json_path = json_path_template.replace("MODULE", list_module[module_count].upper())
                    json_previous_path = json_path_previous_template.replace("MODULE", list_module[module_count].upper())

                    with open(json_previous_path, 'r') as info_prev_file:
                        module_info_prev = json.load(info_prev_file)

                    with open(json_path, 'r') as info_file:
                        module_info = json.load(info_file)

                    for derivetive_count in range(len(self.list_derivative)):
                        prev_data = module_info_prev[self.list_derivative[derivetive_count]]["number_of_test_id_fail"]
                        cur_data  = module_info[self.list_derivative[derivetive_count]]["number_of_test_id_fail"]

                        if cur_data > 0:
                            self.WriteDiffData(prev_data, cur_data, "Failed tests number on ATE", derivetive_count, row_count, color_table.color_list['Red'])
                        else:
                            self.WriteDiffData(prev_data, cur_data, "Failed tests number on ATE", derivetive_count, row_count, group_color)

                        prev_data = module_info_prev[self.list_derivative[derivetive_count]]["number_of_test_id"]
                        cur_data  = module_info[self.list_derivative[derivetive_count]]["number_of_test_id"]
                        self.WriteDiffData(prev_data, cur_data, "Test case number on ATE", derivetive_count, row_count, group_color)

                        prev_data = module_info_prev[self.list_derivative[derivetive_count]]["srs_coverage"]
                        cur_data  = module_info[self.list_derivative[derivetive_count]]["srs_coverage"]
                        if 'yes' in self.is_tm_get_data_from_preliminary:
                            prev_data = module_info_prev[self.list_derivative[derivetive_count]]["preliminary_srs_coverage"]
                            cur_data  = module_info[self.list_derivative[derivetive_count]]["preliminary_srs_coverage"]
                        self.WriteDiffData(prev_data, cur_data, "SRS Coverage Percentage", derivetive_count, row_count, group_color)

                    row_count += 1
            self.list_merge_cell.append(self.RowColToCoordinate(self.col_info["Group"], start_group_row) + ':' + self.RowColToCoordinate(self.col_info["Group"], start_group_row + number_of_module - 1))

            for key in self.addition_header_column.keys():
                for row in range(start_group_row, start_group_row + number_of_module):
                    self.string_to_cell("NULL", self.GetColumnNumber(key), row, group_color)
                    for index in range(1, len(self.addition_header_column[key]["sub_column"])):
                        self.string_to_cell("NULL", self.GetColumnNumber(key) + index, row, group_color)


    def LoadWorkSheet(self):
        self.list_merge_cell = []
        self.active_sheet_name = self.excel_current_sheet_name
        if os.path.exists(self.excel_path):
            self.work_book  = openpyxl.load_workbook(self.excel_path)
            if self.active_sheet_name not in self.work_book.sheetnames:
                self.work_book.create_sheet(title=self.active_sheet_name)
        else:
            self.work_book = openpyxl.Workbook()
            sheet = self.work_book.active
            sheet.title = self.active_sheet_name

        try:
            self.work_book.save(self.excel_path)
        except Exception as error_log:
            print error_log

        self.work_book  = openpyxl.load_workbook(self.excel_path)
        self.work_sheet = self.work_book[self.active_sheet_name]

        self.init_header_column     = self.report_info['init_header_column']
        self.addition_header_column = self.report_info['addition_header_column']

        self.WriteColumnHeader()
        self.ParseJsonToExcel()
        self.makeUpReports()

        try:
            self.work_book.save(self.excel_path)
        except Exception as error_log:
            print error_log

        return self.list_merge_cell