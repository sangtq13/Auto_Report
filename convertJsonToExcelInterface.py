# -*- coding: utf-8 -*-
import os
import sys
from collections import OrderedDict
import json
import shutil
import re
from datetime import datetime
import color_table
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
try:
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter

from configManagmentInterface import ConfigManagmentInterface
class ConvertJsonToExcelInterface(ConfigManagmentInterface):
    """docstring for ConvertJsonToExcelInterface"""
    def __init__(self,p_release_name):
        super(ConvertJsonToExcelInterface, self).__init__(p_release_name)
        current_dir = os.path.abspath(os.path.dirname(__file__))
        self.previos_report_json             = os.path.join(current_dir, ".config" + "\\" + self.release_name + "\\" + self.report_info['previos_report_json'])
        self.current_report_json             = os.path.join(current_dir, ".config" + "\\" + self.release_name + "\\" + self.report_info['current_report_json'])
        self.output_excel_folder             = os.path.join(current_dir, self.report_info['output_excel_folder'])
        self.number_row_header               = self.report_info['number_row_header']
        self.excel_header_row_color          = color_table.color_list[str(''.join(self.report_info['excel_header_row_color']))]
        self.excel_report_name               = self.release_name + self.report_info['excel_report_suffix_name']
        self.list_download_report            = self.report_info['list_download_report']
        self.list_download_derivative        = self.report_info['list_download_derivative']
        self.tm_key_word                     = self.report_info['tm_key_word']
        self.is_tm_get_data_from_preliminary = self.report_info['is_tm_get_data_from_preliminary']
        self.excel_current_sheet_name        = self.report_info['excel_current_sheet_name']
        self.default_column_width            = self.report_info['default_column_width']
        self.group_color                     = self.report_info['group_color']
        self.na_color                        = color_table.color_list[str(''.join(self.report_info['na_color']))]
        self.yes_color                       = color_table.color_list[str(''.join(self.report_info['yes_color']))]
        self.no_color                        = color_table.color_list[str(''.join(self.report_info['no_color']))]
        self.fail_test_sheet_name            = self.report_info['fail_test_sheet_name']
        self.excel_path = os.path.join(self.output_excel_folder, self.excel_report_name)

    def MakeUpdateMergeCell(self, f_list_work_sheet, f_list_merge_cell):
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        self.work_book  = openpyxl.load_workbook(self.excel_path)
        for sheet_index in range(len(f_list_work_sheet)):
            self.work_sheet = self.work_book[f_list_work_sheet[sheet_index]]

            for mergeCell in range(len(f_list_merge_cell[sheet_index])):
                self.style_range(self.work_sheet, str(f_list_merge_cell[sheet_index][mergeCell]), border=border, fill=None, font=None, alignment=None)

        try:
            self.work_book.save(self.excel_path)
        except Exception as error_log:
            print error_log


    def GetColumnHeader(self):
        max_column_header = 0
        list_index        = []

        # a pair of index and column name
        for item in self.init_header_column.keys():
            index = self.init_header_column[item]["column_order"]
            list_index.append(index)
            list_index.append(item)
            if index > max_column_header:
                max_column_header = index

        for item in self.addition_header_column.keys():
            index = self.addition_header_column[item]["column_order"]
            list_index.append(index)
            list_index.append(item)
            if index > max_column_header:
                max_column_header = index

        list_column_header = []
        # index starts from 1
        for count in range(1, max_column_header + 1):
            list_column_header.append(str(list_index[list_index.index(count) + 1]))

        return list_column_header


    def RowColToCoordinate(self, col, row):
        return get_column_letter(col)+ str(row)

    def GetColumnNumber(self, column_name):
        row_search = 1
        for column in range(1, self.work_sheet.max_column + 1):
            if column_name in self.cell_to_string(column, row_search):
                return column


    def cell_to_string(self, f_column, f_row):
        cell_str = str(self.work_sheet.cell(row=f_row, column=f_column).value)

        # filer empty string(spaces, tab, \r\n)
        if cell_str.isspace() or cell_str == 'None':
            return 'N/A'
        else:
            # re.sub() function takes the string argument and replaces one or more space with single space
            # strip() removes all whitespace at the start and end, including spaces, tabs, newlines and carriage returns
            return (re.sub(' +', ' ', cell_str)).strip()


    def string_to_cell(self, f_string, f_column, f_row, f_cell_color, alignment_horizontal='center'):
        cell_value = f_string
        cell_color = f_cell_color
        is_write = True
        if isinstance(f_string, int):
            cell_value = int(f_string)
        if isinstance(f_string, float):
            cell_value = float(f_string)/100

        if isinstance(cell_value, str) or isinstance(cell_value, unicode):
            if 'N/A' in cell_value:
                cell_color = self.na_color
            if 'NULL' in cell_value:
                is_write = False
        current_cell = self.work_sheet.cell(row=f_row, column=f_column)
        if is_write is True:
            current_cell.value = cell_value
        current_cell.font = Font(bold=False)
        current_cell.fill = PatternFill(start_color=cell_color,
                                        end_color=cell_color,
                                        fill_type='solid')

        current_cell.border  = Border(left=Side(style='thin'),
                                      right=Side(style='thin'),
                                      top=Side(style='thin'),
                                      bottom=Side(style='thin'))

        current_cell.alignment = Alignment(wrapText=False,
                                           vertical='center',
                                           horizontal=alignment_horizontal)


    def style_range(self, ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
        """
        Apply styles to a range of cells as if they were a single cell.

        :param ws:  Excel worksheet instance
        :param range: An excel range to style (e.g. A1:F20)
        :param border: An openpyxl Border
        :param fill: An openpyxl PatternFill or GradientFill
        :param font: An openpyxl Font object
        """
        top = Border(top=border.top)
        left = Border(left=border.left)
        right = Border(right=border.right)
        bottom = Border(bottom=border.bottom)

        first_cell = ws[cell_range.split(":")[0]]
        if alignment:
            ws.merge_cells(cell_range)
            first_cell.alignment = alignment

        rows = ws[cell_range]
        if font:
            first_cell.font = font

        for cell in rows[0]:
            cell.border = cell.border + top
        for cell in rows[-1]:
            cell.border = cell.border + bottom

        for row in rows:
            l = row[0]
            r = row[-1]
            l.border = l.border + left
            r.border = r.border + right
            if fill:
                for c in row:
                    c.fill = fill

    def makeUpReports(self):
        # print self.list_merge_cell
        for item in self.list_merge_cell:
            self.work_sheet.merge_cells(item)

        for column_index in range(1, self.work_sheet.max_column + 1):
            self.work_sheet.column_dimensions[get_column_letter(column_index)].width = self.default_column_width

        for key in self.init_header_column.keys():
            if "column_width" in self.init_header_column[key].keys():
                self.work_sheet.column_dimensions[get_column_letter(self.col_info[key])].width = self.init_header_column[key]["column_width"]

        for key in self.addition_header_column.keys():
            if "column_width" in self.addition_header_column[key].keys():
                self.work_sheet.column_dimensions[get_column_letter(self.col_info[key])].width = self.addition_header_column[key]["column_width"]
